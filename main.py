from dateutil import parser
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

typeDescriptionList = ["武器", "角色",
                       "Weapon", "Character",
                       "武器", "キャラクター"]


def typeChecker(stringRecord):
    try:
        typeResult = parser.parse(stringRecord)
    except:
        for typeDescription in typeDescriptionList:
            if typeDescription in stringRecord:
                typeResult = "type"
                break
        else:
            typeResult = "name"
    return typeResult


def txt2df(txtInput):
    txtFile = open(txtInput, 'r', encoding='UTF-8')
    fullSet = []
    currentSet = ["", "", ""]
    for line in txtFile:
        result = typeChecker(line)
        if result == "type":
            currentSet[0] = line
        elif result == "name":
            currentSet[1] = line
        else:
            currentSet[2] = result
            if currentSet[1] != [""]:
                fullSet.append(currentSet)
            currentSet = ["", "", ""]
    # print(fullSet)
    txtFile.close()
    df = pd.DataFrame(fullSet)
    return df


def excelReformat(wishType):
    wb = load_workbook("GenshinInpactWishHistory.xlsx")
    ws = wb[wishType]

    # 高亮四星五星出货
    for cell in ws["B"]:
        if "五星" in cell.value:
            # print(cell.value)
            cell.font = Font(color="00FF9900", bold=True)
        if "四星" in cell.value:
            # print(cell.value)
            cell.font = Font(color="00FF00FF", bold=True)

    # 添加标题
    ws.insert_rows(0)
    if wishType == "Novice":
        title = "新手祈愿"
    elif wishType == "Permanent":
        title = "常驻祈愿"
    elif wishType == "CharacterEvent":
        title = "角色活动UP池"
    elif wishType == "WeaponEvent":
        title = "武器活动UP池"
    ws['A1'].value = title
    ws['A1'].font = Font(bold=True)  # 加粗
    ws['A1'].alignment = Alignment(horizontal='center')  # 居中
    ws['A1'].fill = PatternFill(start_color="0099CCFF", fill_type="solid")  # 填充颜色
    ws.merge_cells('A1:C1')  # 合并单元格

    # 设置Column Width
    ws.column_dimensions['A'].width = 5.15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 21

    # 保存文件
    wb.save("GenshinInpactWishHistory.xlsx")


if __name__ == "__main__":
    print("原神祈愿历史记录导出工具  Genshin Impact Wish History Export Utility")
    print("https://github.com/Masterain98/GenshinImpactWishHistoryUtility")
    print(" ")
    print("请按序号输入需要导出的数据：")
    print("1. 新手祈愿")
    print("2. 常驻祈愿")
    print("3. 角色活动UP")
    print("4. 武器活动UP")
    userInput = input("请输入序号并以空格间隔：")
    if "1" in userInput:
        Novice = txt2df("Novice.txt")
    if "2" in userInput:
        Permanent = txt2df("Permanent.txt")
    if "3" in userInput:
        CharacterEvent = txt2df("CharacterEvent.txt")
    if "4" in userInput:
        WeaponEvent = txt2df("WeaponEvent.txt")

    with pd.ExcelWriter('GenshinInpactWishHistory.xlsx', engine='openpyxl') as writer:
        if "1" in userInput:
            Novice.to_excel(writer, sheet_name='Novice', index=False, header=["种类", "名称", "时间"])
        if "2" in userInput:
            Permanent.to_excel(writer, sheet_name='Permanent', index=False, header=["种类", "名称", "时间"])
        if "3" in userInput:
            CharacterEvent.to_excel(writer, sheet_name='CharacterEvent', index=False, header=["种类", "名称", "时间"])
        if "4" in userInput:
            WeaponEvent.to_excel(writer, sheet_name='WeaponEvent', index=False, header=["种类", "名称", "时间"])
    writer.save()
    if "1" in userInput:
        excelReformat("Novice")
    if "2" in userInput:
        excelReformat("Permanent")
    if "3" in userInput:
        excelReformat("CharacterEvent")
    if "4" in userInput:
        excelReformat("WeaponEvent")
    input("导出完成！按任意键退出...")
